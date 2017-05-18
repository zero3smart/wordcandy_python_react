from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.generics import GenericAPIView
from rest_framework import status

from django.http import HttpResponse
from django.conf import settings
from django.core.cache import cache

import time
import requests
from datetime import datetime

from .serializers import SynonymsSerializer, AntonymsSerializer, ShopSerializer, TemplateSerializer, SubscribeSerializer, ExportSerializer
from .models import Shop, Subscribe, Word

from openpyxl import Workbook
from openpyxl.drawing.image import Image

from rest_framework_tracking.mixins import LoggingMixin


class SynonymsView(LoggingMixin, GenericAPIView):

    def get(self, request, format=None):
        """
        Return words with synonyms
        """
        data = {}
        data['synonyms'] = []
        tags = request.GET.get('tags', '')
        for tag in tags.split(','):
            result = requests.get('https://wordsapiv1.p.mashape.com/words/{0}/synonyms'.format(tag),
                                  headers={
                "X-Mashape-Key": settings.MASHAPE,
                "Accept": "application/json",
            })
            if 'synonyms' in result.json():
                data['synonyms'] += result.json()['synonyms']
        return Response(data)


class AntonymsView(LoggingMixin, GenericAPIView):

    def get(self, request, format=None):
        """
        Return words with antonyms
        """
        data = {}
        data['antonyms'] = []
        tags = request.GET.get('tags', '')
        for tag in tags.split(','):
            result = requests.get('https://wordsapiv1.p.mashape.com/words/{0}/antonyms'.format(tag),
                                  headers={
                "X-Mashape-Key": settings.MASHAPE,
                "Accept": "application/json",
            })
            if 'antonyms' in result.json():
                data['antonyms'] += result.json()['antonyms']
        return Response(data)


class KeywordToolView(LoggingMixin, GenericAPIView):

    def get(self, request, format=None):
        """
        Return result from keywordtool
        """
        keywords = request.GET.get('tags', '')
        result = {
            'keywords': []
        }

        if keywords != '':
            for word in keywords.split(','):
                payload = {
                    'apikey': settings.KEYWORDTOOL,
                    'keyword': '[{0}]'.format(word),
                    'output': 'json',
                    'country': 'us',
                    'language': 'en',
                    'metrics': 'true',
                    'metrics_location': '2840',
                    'metrics_language': 'en'
                }
                word = word.lower()

                word_result = Word.objects.filter(name=word).first()

                if word_result:
                    data = word_result.results
                else:
                    try:
                        data_keywordtool = requests.get('http://api.keywordtool.io/v2/search/suggestions/amazon', params=payload)
                        results = data_keywordtool.json()
                        created_word = Word.objects.create(name=word, results=results)
                        data = created_word.results
                    except Exception as e:
                        data = False

                if data:
                    for item in data['results']:
                        for sub_item in data['results'][item]:
                            if 'volume' in sub_item and 'string' in sub_item:
                                result['keywords'].append({'name': sub_item['string'], 'volume': sub_item['volume']})

        return Response(result)


class ExcelView(LoggingMixin, GenericAPIView):
    serializer_class = ExportSerializer

    def post(self, request, format=None):
        """
        Return excel file
        """
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save()
            wb = Workbook()
            ws = wb.active
            ws['A1'] = 'TEE SHIRT DESIGN'
            if serializer.data['photo']:
                img = Image('{0}{1}'.format(settings.MEDIA_ROOT[:-6], serializer.data['photo']))
                ws.add_image(img, 'A2')
            ws['B1'] = 'PRODUCT TITLE'
            ws['B2'] = serializer.data['product_name']
            ws['C1'] = 'BULLET POINT ONE'
            ws['C2'] = serializer.data['first_description']
            ws['D1'] = 'BULET POINT TWO'
            ws['D2'] = serializer.data['second_description']
            ws['E1'] = 'SELECTED KEYWORDS FROM WC'
            ws['E2'] = serializer.data['keywords']
            timestamp = int(time.time())
            wb.save('{0}/exel/{1}.xlsx'.format(settings.MEDIA_ROOT, timestamp))
            result = {
                'file': '{0}{1}exel/{2}.xlsx'.format(settings.WEBSITE, settings.MEDIA_URL, timestamp),
                'data': serializer.data
            }
            return Response(result, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ShopList(LoggingMixin, GenericAPIView):
    serializer_class = ShopSerializer

    def get(self, request, format=None):
        """
        Return list of shops
        """
        shops = Shop.objects.all()
        serializer = self.serializer_class(shops, many=True)
        return Response(serializer.data)


class SubscribeView(LoggingMixin, GenericAPIView):
    serializer_class = SubscribeSerializer

    def post(self, request, format=None):
        """
        Return add email to subscribe
        """
        serializer = self.serializer_class(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)